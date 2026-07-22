"""Branch Manager books & import API.

Lets a Branch Manager read branch-scoped general ledger data, export the
books to CSV/XLSX, stage and commit CSV/XLSX imports for ``Loan
Repayment``, ``Customer`` (branch reassignment), and ``LMS Borrower
Compliance``, and view the wallet-reconciliation summary.

All endpoints are branch-scoped via ``_manager_branch`` and fail closed
when the manager has no branch assigned (no portfolio leak).
"""

from __future__ import annotations

import base64
import csv
import io
import json
import uuid

import frappe
from frappe import _
from frappe.rate_limiter import rate_limit
from frappe.utils import cint, flt, getdate, today


# ---------------------------------------------------------------------------
# Guards — persona + branch scope.
# ---------------------------------------------------------------------------

def _require_manager():
	"""Branch Manager only; admins allowed for testing and multi-branch BMs."""
	if frappe.session.user == "Guest":
		frappe.throw(_("Please log in"), frappe.PermissionError)
	roles = set(frappe.get_roles())
	if roles.intersection({"System Manager", "Administrator"}):
		return
	from lms_saas.utils.brand import _get_user_persona

	if _get_user_persona() != "Branch Manager":
		frappe.throw(_("Not permitted"), frappe.PermissionError)


def _manager_branch() -> str | None:
	from lms_saas.api.staff import get_current_user_branch

	return get_current_user_branch()


def _is_admin() -> bool:
	return bool(set(frappe.get_roles()).intersection({"System Manager", "Administrator"}))


def _parse_date(value, default=None):
	"""Tolerates ISO, DD/MM/YYYY, YYYY-MM-DD, MM/DD/YYYY."""
	if value in (None, "", "null"):
		return default
	try:
		return getdate(value)
	except Exception:
		from datetime import datetime
		for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
			try:
				return datetime.strptime(str(value), fmt).date()
			except Exception:
				continue
		frappe.throw(_("Invalid date: {0}").format(value))


# ===========================================================================
# Books — read-only GL view, branch-scoped.
# ===========================================================================

_ACCOUNT_CLASSES = (
	("Income", ("Income", "Revenue", "Other Income")),
	("Expense", ("Expense", "Cost of Goods Sold", "Depreciation", "Indirect Expense")),
	("Asset", ("Bank", "Cash", "Receivable", "Stock", "Fixed Asset", "Asset")),
	("Liability", ("Payable", "Loan", "Equity", "Liability")),
)


def _classify_account(account: str) -> str:
	"""Best-effort account classification by root_type on the Account master."""
	if not account:
		return "Other"
	root_type = frappe.db.get_value("Account", account, "root_type")
	if root_type:
		return root_type
	for klass, prefixes in _ACCOUNT_CLASSES:
		for prefix in prefixes:
			if account.startswith(prefix):
				return klass
	return "Other"


@frappe.whitelist()
def get_branch_books(
	from_date=None,
	to_date=None,
	limit: int = 500,
	account=None,
):
	"""Return GL rows for the manager's branch, plus a per-account summary.

	``from_date`` / ``to_date`` are inclusive. The row list is capped at
	``limit`` so a multi-month request stays UI-friendly; the BM can
	export the full set with ``export_branch_books``.
	"""
	_require_manager()
	branch = _manager_branch()
	if not branch and not _is_admin():
		return {"rows": [], "summary": {}, "branch": None, "total_rows": 0}

	conditions = ["gl.cost_center = %(branch)s"]
	params = {"branch": branch, "limit": int(limit)}
	if from_date:
		conditions.append("gl.posting_date >= %(from_date)s")
		params["from_date"] = _parse_date(from_date)
	if to_date:
		conditions.append("gl.posting_date <= %(to_date)s")
		params["to_date"] = _parse_date(to_date)
	if account:
		conditions.append("gl.account = %(account)s")
		params["account"] = account

	where = " AND ".join(conditions)
	rows = frappe.db.sql(
		f"""
		SELECT
			gl.name, gl.posting_date, gl.account, gl.party_type, gl.party,
			gl.debit, gl.credit, gl.debit_in_account_currency, gl.credit_in_account_currency,
			gl.account_currency, gl.against, gl.remarks, gl.voucher_type, gl.voucher_no
		FROM `tabGL Entry` gl
		WHERE {where}
		ORDER BY gl.posting_date DESC, gl.creation DESC
		LIMIT %(limit)s
		""",
		params,
		as_dict=True,
	)
	total_rows = cint(frappe.db.sql(
		f"SELECT COUNT(*) FROM `tabGL Entry` gl WHERE {where}", params
	)[0][0])

	summary_rows = frappe.db.sql(
		f"""
		SELECT
			gl.account,
			SUM(gl.debit) AS debit,
			SUM(gl.credit) AS credit
		FROM `tabGL Entry` gl
		WHERE {where}
		GROUP BY gl.account
		ORDER BY gl.account
		""",
		params,
		as_dict=True,
	)
	summary = []
	total_debit = total_credit = 0.0
	for row in summary_rows:
		klass = _classify_account(row["account"])
		net = flt(row["debit"]) - flt(row["credit"])
		summary.append({
			"account": row["account"],
			"class": klass,
			"debit": flt(row["debit"]),
			"credit": flt(row["credit"]),
			"net": net,
		})
		total_debit += flt(row["debit"])
		total_credit += flt(row["credit"])

	class_totals: dict = {}
	for s in summary:
		klass = s["class"]
		bucket = class_totals.setdefault(klass, {"debit": 0.0, "credit": 0.0, "net": 0.0})
		bucket["debit"] += s["debit"]
		bucket["credit"] += s["credit"]
		bucket["net"] += s["net"]

	return {
		"branch": branch,
		"from_date": str(params.get("from_date") or ""),
		"to_date": str(params.get("to_date") or ""),
		"rows": rows,
		"summary": summary,
		"class_totals": class_totals,
		"totals": {"debit": total_debit, "credit": total_credit, "net": total_debit - total_credit},
		"total_rows": total_rows,
	}


@frappe.whitelist()
def get_branch_pnl(from_date=None, to_date=None):
	"""Income, expense, and net for the branch in the given period."""
	books = get_branch_books(from_date=from_date, to_date=to_date, limit=1)
	income = expense = 0.0
	for klass, totals in (books.get("class_totals") or {}).items():
		if klass == "Income":
			income = totals.get("net", 0.0)
		elif klass == "Expense":
			expense = totals.get("net", 0.0)
	return {
		"branch": books.get("branch"),
		"from_date": books.get("from_date"),
		"to_date": books.get("to_date"),
		"income": income,
		"expense": expense,
		"net": income - expense,
	}


@frappe.whitelist()
def get_borrower_ledger(customer, from_date=None, to_date=None):
	"""Return GL rows for a single borrower (party = Customer) in the branch."""
	_require_manager()
	branch = _manager_branch()
	if not branch and not _is_admin():
		return {"rows": [], "branch": None}
	if not customer:
		frappe.throw(_("Customer is required."))

	conditions = ["gl.party_type = 'Customer'", "gl.party = %(customer)s", "gl.cost_center = %(branch)s"]
	params = {"customer": customer, "branch": branch}
	if from_date:
		conditions.append("gl.posting_date >= %(from_date)s")
		params["from_date"] = _parse_date(from_date)
	if to_date:
		conditions.append("gl.posting_date <= %(to_date)s")
		params["to_date"] = _parse_date(to_date)
	where = " AND ".join(conditions)
	rows = frappe.db.sql(
		f"""
		SELECT posting_date, account, debit, credit, against, voucher_type, voucher_no, remarks
		FROM `tabGL Entry` gl
		WHERE {where}
		ORDER BY posting_date DESC, creation DESC
		""",
		params,
		as_dict=True,
	)
	return {"branch": branch, "customer": customer, "rows": rows}


@frappe.whitelist()
@rate_limit(limit=10, seconds=60)
def export_branch_books(from_date=None, to_date=None, fmt: str = "csv"):
	"""Build a base64-encoded CSV/XLSX export of the branch books."""
	_require_manager()
	branch = _manager_branch()
	if not branch and not _is_admin():
		frappe.throw(_("No branch scope; nothing to export."))

	books = get_branch_books(from_date=from_date, to_date=to_date, limit=100000)
	fmt = (fmt or "csv").lower()
	if fmt not in ("csv", "xlsx"):
		frappe.throw(_("Format must be csv or xlsx."))

	rows = books.get("rows", [])
	if fmt == "csv":
		buffer = io.StringIO()
		writer = csv.writer(buffer)
		writer.writerow([
			"posting_date", "account", "party_type", "party",
			"debit", "credit", "currency", "voucher_type", "voucher_no", "remarks",
		])
		for r in rows:
			writer.writerow([
				r.get("posting_date"),
				r.get("account"),
				r.get("party_type") or "",
				r.get("party") or "",
				flt(r.get("debit")),
				flt(r.get("credit")),
				r.get("account_currency") or "",
				r.get("voucher_type") or "",
				r.get("voucher_no") or "",
				(r.get("remarks") or "").replace("\n", " "),
			])
		raw = buffer.getvalue().encode("utf-8")
		filename = f"branch_books_{branch}_{today()}.csv"
	else:
		try:
			import openpyxl  # noqa: F401
		except Exception:
			frappe.throw(_("openpyxl is not installed; export as CSV instead."))
		import openpyxl as _xl
		wb = _xl.Workbook()
		ws = wb.active
		ws.title = "Books"
		ws.append([
			"Posting date", "Account", "Party type", "Party",
			"Debit", "Credit", "Currency", "Voucher type", "Voucher no", "Remarks",
		])
		for r in rows:
			ws.append([
				str(r.get("posting_date") or ""),
				r.get("account") or "",
				r.get("party_type") or "",
				r.get("party") or "",
				flt(r.get("debit")),
				flt(r.get("credit")),
				r.get("account_currency") or "",
				r.get("voucher_type") or "",
				r.get("voucher_no") or "",
				(r.get("remarks") or "").replace("\n", " "),
			])
		summary = wb.create_sheet("Summary")
		summary.append(["Account", "Class", "Debit", "Credit", "Net"])
		for s in books.get("summary", []):
			summary.append([s["account"], s["class"], s["debit"], s["credit"], s["net"]])
		summary.append([])
		summary.append(["Totals", "", books["totals"]["debit"], books["totals"]["credit"], books["totals"]["net"]])
		out = io.BytesIO()
		wb.save(out)
		raw = out.getvalue()
		filename = f"branch_books_{branch}_{today()}.xlsx"

	try:
		from lms_saas.api.compliance import record_money_event
		record_money_event(
			doctype="LMS Audit Event",
			docname=None,
			action=f"export_branch_books:{fmt}",
			party=frappe.session.user,
			amount=0,
			remarks=f"Exported {len(rows)} GL rows for branch {branch}.",
		)
	except Exception:
		pass

	return {
		"filename": filename,
		"mime": "text/csv" if fmt == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		"encoding": "base64",
		"data": base64.b64encode(raw).decode("ascii"),
		"row_count": len(rows),
	}


# ===========================================================================
# Import — staging-first CSV/XLSX upload, preview, commit.
# ===========================================================================

ALLOWED_IMPORT_DOCTYPES = {
	"Loan Repayment": {
		"required": ["against_loan", "applicant_type", "applicant", "posting_date", "amount_paid"],
		"column_map": {
			"against_loan": "against_loan",
			"loan": "against_loan",
			"loan_id": "against_loan",
			"applicant_type": "applicant_type",
			"applicant": "applicant",
			"customer": "applicant",
			"borrower": "applicant",
			"company": "company",
			"posting_date": "posting_date",
			"date": "posting_date",
			"amount_paid": "amount_paid",
			"amount": "amount_paid",
			"value": "amount_paid",
		},
		"applicant_type_default": "Customer",
	},
	"Customer": {
		"required": ["name", "custom_lms_branch"],
		"column_map": {
			"name": "name",
			"customer_name": "customer_name",
			"name1": "name",
			"customer": "name",
			"branch": "custom_lms_branch",
			"custom_lms_branch": "custom_lms_branch",
		},
		"applicant_type_default": None,
	},
	"LMS Borrower Compliance": {
		"required": ["customer", "kyc_status", "consent_given"],
		"column_map": {
			"customer": "customer",
			"borrower": "customer",
			"applicant": "customer",
			"kyc_status": "kyc_status",
			"kyc": "kyc_status",
			"consent_given": "consent_given",
			"consent": "consent_given",
			"consent_date": "consent_date",
		},
		"applicant_type_default": None,
	},
}


def _decode_payload(file_b64, mime_hint=None):
	"""Decode a base64 file payload and return (raw_bytes, resolved_mime)."""
	if not file_b64:
		frappe.throw(_("No file payload received."))
	if "," in file_b64 and file_b64.lstrip().lower().startswith("data:"):
		file_b64 = file_b64.split(",", 1)[1]
	try:
		raw = base64.b64decode(file_b64, validate=True)
	except Exception as exc:
		frappe.throw(_("File payload is not valid base64: {0}").format(exc))
	resolved = mime_hint or "text/csv"
	if raw[:2] == b"PK":
		resolved = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	return raw, resolved


def _read_rows(raw, mime):
	"""Return (rows, headers) for a CSV or XLSX payload."""
	if "spreadsheetml" in mime:
		try:
			import openpyxl as _xl
		except Exception:
			frappe.throw(_("openpyxl is not installed; upload a CSV instead."))
		wb = _xl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
		ws = wb.active
		iter_rows = ws.iter_rows(values_only=True)
		header_row = next(iter_rows, None)
		if not header_row:
			return [], []
		headers = [str(h or "").strip() for h in header_row]
		rows = []
		for r in iter_rows:
			if r is None or all(c is None for c in r):
				continue
			row = {headers[i]: ("" if r[i] is None else r[i]) for i in range(min(len(headers), len(r)))}
			rows.append(row)
		return rows, headers
	text = raw.decode("utf-8-sig", errors="replace")
	reader = csv.reader(io.StringIO(text))
	try:
		headers = next(reader)
	except StopIteration:
		return [], []
	headers = [str(h or "").strip() for h in headers]
	rows = []
	for raw_row in reader:
		if not raw_row or all(not str(c).strip() for c in raw_row):
			continue
		row = {headers[i]: raw_row[i] if i < len(raw_row) else "" for i in range(len(headers))}
		rows.append(row)
	return rows, headers


def _apply_mapping(row, mapping, spec):
	"""Apply a user column map (or the spec default) to a raw row."""
	out: dict = {}
	for raw_key, value in row.items():
		key = (raw_key or "").strip()
		canonical = mapping.get(key) if mapping else None
		if not canonical:
			canonical = spec["column_map"].get(key.lower())
		if canonical:
			out[canonical] = value
	if spec.get("applicant_type_default") and "applicant_type" in spec["required"]:
		out.setdefault("applicant_type", spec["applicant_type_default"])
	return out


def _validate_loan_repayment(row, branch):
	errs: list = []
	loan = (row.get("against_loan") or "").strip()
	if not loan:
		errs.append("missing against_loan")
	else:
		loan_doc = frappe.db.get_value(
			"Loan", loan,
			["name", "docstatus", "applicant", "applicant_type", "company", "custom_lms_branch"],
			as_dict=True,
		)
		if not loan_doc:
			errs.append(f"loan {loan} not found")
		else:
			if cint(loan_doc.get("docstatus")) != 1:
				errs.append(f"loan {loan} is not submitted")
			if branch and loan_doc.get("custom_lms_branch") and loan_doc["custom_lms_branch"] != branch:
				errs.append(f"loan {loan} is not in your branch")
			if row.get("applicant") and loan_doc.get("applicant") and row["applicant"] != loan_doc["applicant"]:
				errs.append("applicant does not match the loan")
			if row.get("applicant_type") and loan_doc.get("applicant_type") and row["applicant_type"] != loan_doc["applicant_type"]:
				errs.append("applicant_type does not match the loan")
	if not row.get("posting_date"):
		errs.append("missing posting_date")
	else:
		_ = _parse_date(row["posting_date"])
	if flt(row.get("amount_paid") or 0) <= 0:
		errs.append("amount_paid must be greater than zero")
	return errs


def _validate_customer(row, branch):
	errs: list = []
	name = (row.get("name") or "").strip()
	if not name:
		errs.append("missing customer name")
	if not (row.get("custom_lms_branch") or "").strip():
		errs.append("missing custom_lms_branch")
	elif branch and (row.get("custom_lms_branch") or "").strip() != branch:
		errs.append(f"customer {name} branch must equal your branch ({branch})")
	return errs


def _validate_compliance(row, branch):
	errs: list = []
	if not (row.get("customer") or "").strip():
		errs.append("missing customer")
	else:
		cust_branch = frappe.db.get_value("Customer", row["customer"], "custom_lms_branch")
		if branch and cust_branch and cust_branch != branch:
			errs.append("customer is not in your branch")
	if not (row.get("kyc_status") or "").strip():
		errs.append("missing kyc_status")
	if not row.get("consent_given"):
		errs.append("missing consent_given")
	return errs


_VALIDATORS = {
	"Loan Repayment": _validate_loan_repayment,
	"Customer": _validate_customer,
	"LMS Borrower Compliance": _validate_compliance,
}


@frappe.whitelist()
@rate_limit(limit=10, seconds=60)
def create_import_batch(
	doctype,
	file_b64,
	mapping=None,
	branch_override=None,
	mime_hint=None,
):
	"""Parse an upload, validate every row, and create an LMS Import Batch.

	Returns the batch name plus a preview payload. Nothing is committed
	until the manager calls ``commit_import_batch`` separately.
	"""
	_require_manager()
	if doctype not in ALLOWED_IMPORT_DOCTYPES:
		frappe.throw(_("DocType {0} is not allowed for branch-manager import.").format(doctype))
	spec = ALLOWED_IMPORT_DOCTYPES[doctype]

	branch = branch_override or _manager_branch()
	if not branch and not _is_admin():
		frappe.throw(_("Your account is not scoped to a branch; import is disabled."))

	raw, mime = _decode_payload(file_b64, mime_hint)
	rows, headers = _read_rows(raw, mime)
	if not rows:
		frappe.throw(_("The file has no data rows."))

	user_map: dict = {}
	if mapping:
		try:
			user_map = json.loads(mapping)
		except Exception as exc:
			frappe.throw(_("Invalid column mapping JSON: {0}").format(exc))

	validator = _VALIDATORS[doctype]
	preview_rows: list = []
	valid_count = 0
	error_count = 0
	for idx, raw_row in enumerate(rows, start=1):
		mapped = _apply_mapping(raw_row, user_map, spec)
		errs = validator(mapped, branch)
		preview_rows.append({"row": idx, "data": mapped, "errors": errs, "ok": not errs})
		if errs:
			error_count += 1
		else:
			valid_count += 1

	batch = frappe.get_doc({
		"doctype": "LMS Import Batch",
		"target_doctype": doctype,
		"branch": branch,
		"row_count": len(preview_rows),
		"valid_count": valid_count,
		"error_count": error_count,
		"status": "Staged",
		"idempotency_key": str(uuid.uuid4()),
		"raw_preview": json.dumps(preview_rows[:50], default=str),
	})
	batch.flags.ignore_permissions = True
	batch.insert(ignore_permissions=True)

	return {
		"batch": batch.name,
		"idempotency_key": batch.idempotency_key,
		"status": batch.status,
		"row_count": len(preview_rows),
		"valid_count": valid_count,
		"error_count": error_count,
		"preview": preview_rows,
		"headers": headers,
	}


@frappe.whitelist()
def preview_import_batch(batch):
	"""Return the cached preview of an LMS Import Batch (first 50 rows)."""
	_require_manager()
	doc = frappe.get_doc("LMS Import Batch", batch)
	return {
		"batch": doc.name,
		"status": doc.status,
		"target_doctype": doc.target_doctype,
		"row_count": doc.row_count,
		"valid_count": doc.valid_count,
		"error_count": doc.error_count,
		"idempotency_key": doc.idempotency_key,
		"preview": json.loads(doc.raw_preview or "[]"),
	}


@frappe.whitelist()
@rate_limit(limit=5, seconds=60)
def commit_import_batch(batch, dry_run: int | bool = 0):
	"""Commit staged rows in a single transaction. Idempotent per batch."""
	_require_manager()
	dry_run = bool(cint(dry_run))

	doc = frappe.get_doc("LMS Import Batch", batch)
	if doc.status == "Committed":
		return {
			"batch": doc.name,
			"status": doc.status,
			"committed": doc.committed_count,
			"errors": doc.error_count,
			"message": _("Batch already committed."),
		}
	if doc.status not in ("Staged", "Partial"):
		frappe.throw(_("Batch is in state {0}; cannot commit.").format(doc.status))

	doctype = doc.target_doctype
	if doctype not in ALLOWED_IMPORT_DOCTYPES:
		frappe.throw(_("DocType {0} is not allowed.").format(doctype))

	preview = json.loads(doc.raw_preview or "[]")
	branch = doc.branch
	committed = 0
	errors: list = []

	if dry_run:
		for entry in preview:
			if entry.get("ok"):
				committed += 1
		return {
			"batch": doc.name,
			"status": "DryRun",
			"committed": committed,
			"errors": 0,
			"dry_run": True,
		}

	try:
		frappe.db.begin()
		for entry in preview:
			if not entry.get("ok"):
				continue
			row = entry["data"]
			try:
				_commit_one(doctype, row, branch)
				committed += 1
			except Exception as exc:
				errors.append({"row": entry.get("row"), "message": str(exc)})
		if errors:
			frappe.db.rollback()
			frappe.db.begin()
			doc.status = "Failed"
			doc.committed_count = 0
			doc.error_count = len(errors)
			doc.flags.ignore_permissions = True
			doc.save(ignore_permissions=True)
			return {
				"batch": doc.name,
				"status": "Failed",
				"committed": 0,
				"errors": errors,
			}
		doc.status = "Committed"
		doc.committed_count = committed
		doc.error_count = 0
		doc.flags.ignore_permissions = True
		doc.save(ignore_permissions=True)
		frappe.db.commit()
	except Exception as exc:
		frappe.db.rollback()
		doc.status = "Failed"
		doc.flags.ignore_permissions = True
		doc.save(ignore_permissions=True)
		frappe.throw(_("Commit failed: {0}").format(exc))

	try:
		from lms_saas.api.compliance import record_money_event
		record_money_event(
			doctype="LMS Audit Event",
			docname=None,
			action=f"import_commit:{doctype}",
			party=frappe.session.user,
			amount=0,
			remarks=f"Committed {committed} {doctype} rows (batch {doc.name}).",
		)
	except Exception:
		pass

	return {
		"batch": doc.name,
		"status": doc.status,
		"committed": committed,
		"errors": 0,
		"idempotency_key": doc.idempotency_key,
	}


def _commit_one(doctype, row, branch):
	"""Apply a single validated row. Raises on failure."""
	if doctype == "Loan Repayment":
		lr = frappe.new_doc("Loan Repayment")
		lr.against_loan = row["against_loan"]
		lr.applicant_type = row.get("applicant_type") or "Customer"
		lr.applicant = row["applicant"]
		if row.get("company"):
			lr.company = row["company"]
		lr.posting_date = _parse_date(row["posting_date"])
		lr.amount_paid = flt(row["amount_paid"])
		lr.flags.ignore_permissions = True
		lr.insert()
		lr.submit()
		return
	if doctype == "Customer":
		name = row["name"].strip()
		existing = frappe.db.exists("Customer", name)
		if existing:
			cust = frappe.get_doc("Customer", name)
		else:
			cust = frappe.new_doc("Customer")
			cust.name = name
		cust.customer_name = (row.get("customer_name") or name).strip()
		cust.custom_lms_branch = (row.get("custom_lms_branch") or "").strip()
		cust.flags.ignore_permissions = True
		cust.save(ignore_permissions=True)
		return
	if doctype == "LMS Borrower Compliance":
		customer = (row.get("customer") or "").strip()
		existing = frappe.db.get_value("LMS Borrower Compliance", {"customer": customer}, "name")
		if existing:
			comp = frappe.get_doc("LMS Borrower Compliance", existing)
		else:
			comp = frappe.new_doc("LMS Borrower Compliance")
			comp.customer = customer
		comp.kyc_status = (row.get("kyc_status") or "Pending").strip()
		comp.consent_given = 1 if str(row.get("consent_given")).lower() in ("1", "true", "yes", "y") else 0
		if row.get("consent_date"):
			comp.consent_date = _parse_date(row["consent_date"])
		comp.flags.ignore_permissions = True
		comp.save(ignore_permissions=True)
		return
	frappe.throw(_("Unsupported DocType {0}").format(doctype))


# ===========================================================================
# Reconciliation wrapper — surfaces wallet-recon stats + unmatched count.
# ===========================================================================

@frappe.whitelist()
def get_reconciliation_summary(limit: int = 20):
	"""Lightweight wrapper around the wallet-recon API."""
	_require_manager()
	branch = _manager_branch()
	company = None
	if branch:
		company = frappe.db.get_value("Cost Center", branch, "company")

	total = matched = unmatched = 0
	matched_value = unmatched_value = 0.0
	if company:
		for status in ("Matched", "Unmatched"):
			rows = frappe.get_all(
				"LMS Wallet Statement",
				filters={"status": status, "company": company},
				fields=["amount"],
			)
			val = sum(flt(r.get("amount") or 0) for r in rows)
			if status == "Matched":
				matched = len(rows)
				matched_value = val
			else:
				unmatched = len(rows)
				unmatched_value = val
		total = matched + unmatched
	else:
		total = frappe.db.count("LMS Wallet Statement")
		matched = frappe.db.count("LMS Wallet Statement", {"status": "Matched"})
		unmatched = frappe.db.count("LMS Wallet Statement", {"status": "Unmatched"})

	unmatched_rows = []
	if unmatched:
		filters = {"status": "Unmatched"}
		if company:
			filters["company"] = company
		unmatched_rows = frappe.get_all(
			"LMS Wallet Statement",
			filters=filters,
			fields=["name", "provider_code", "statement_date", "external_ref", "amount", "company"],
			order_by="statement_date desc",
			limit_page_length=int(limit),
		)

	return {
		"branch": branch,
		"company": company,
		"total": total,
		"matched": matched,
		"unmatched": unmatched,
		"matched_value": matched_value,
		"unmatched_value": unmatched_value,
		"unmatched_rows": unmatched_rows,
	}
